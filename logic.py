import os
import pandas as pd
import pycountry
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from config import AppConfig

class DataConverter:
    def __init__(self, log_callback=None):
        self.log_callback = log_callback

    def log(self, msg):
        if self.log_callback: self.log_callback(msg)
        else: print(msg)

    def run(self, input_path, output_path, filter_mode="Alle"):
        self.log(f"Backend: Starte Job (Filter: {filter_mode})...")
        try:
            df = self._read_csv(input_path)
            df = self._apply_status_filter(df, filter_mode)

            cols_preview = ", ".join(list(df.columns)[:5]) + "..."
            self.log(f"Spalten: {cols_preview}")

            df_export = self._map_columns(df)
            df_export = self._clean_data(df_export)
            df_export = self._sort_data(df_export)

            self._save_excel(df_export, output_path)
            self.log("Backend: Fertig.")
            return True
        except Exception as e:
            self.log(f"FEHLER: {e}")
            raise e

    def _read_csv(self, path):
        try: return pd.read_csv(path, encoding='utf-8')
        except:
            try: return pd.read_csv(path, encoding='latin1')
            except: return pd.read_csv(path, sep=';', encoding='latin1')

    def _apply_status_filter(self, df, mode):
        df.columns = [str(c).strip() for c in df.columns]

        status_col = None
        for col in df.columns:
            for search in AppConfig.STATUS_COL_SEARCH:
                if search in col.lower():
                    status_col = col
                    break
            if status_col: break

        if not status_col:
            self.log("WARNUNG: Keine Status-Spalte gefunden! Filter wird übersprungen.")
            return df

        df = df[~df[status_col].astype(str).str.lower().isin(AppConfig.STATUS_IGNORE_VALUES)]

        if mode == "Nur Bezahlt":
            df = df[df[status_col].astype(str).str.lower().isin(AppConfig.STATUS_PAID_VALUES)]
        elif mode == "Nur Offen":
            df = df[~df[status_col].astype(str).str.lower().isin(AppConfig.STATUS_PAID_VALUES)]

        return df

    def _map_columns(self, df):
        df_export = pd.DataFrame()
        for ziel in AppConfig.ZIEL_SPALTEN:
            such = AppConfig.SUCH_MAPPING.get(ziel, [])
            gefunden = None
            for col in df.columns:
                for begriff in such:
                    if begriff.lower() in str(col).lower():
                        gefunden = col
                        break
                if gefunden: break

            if gefunden:
                df_export[ziel] = df[gefunden]
            else:
                df_export[ziel] = ""
        return df_export

    def _clean_data(self, df):
        df = df.fillna('')
        try: df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
        except: df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

        for col in AppConfig.SMART_CASING_COLUMNS:
            if col in df.columns:
                df[col] = df[col].astype(str).apply(lambda x: x.title() if x else x)

        if 'Your country:' in df.columns:
            df['Your country:'] = df['Your country:'].apply(self._get_country_name)

        if 'City' in df.columns:
            df['City'] = df['City'].apply(self._clean_city_name)

        return df

    def _get_country_name(self, value):
        if not value: return ""
        text = str(value).strip().upper()
        if len(text) == 2:
            try:
                country = pycountry.countries.get(alpha_2=text)
                if country: return getattr(country, 'common_name', country.name)
            except: pass
        return str(value)

    def _clean_city_name(self, value):
        if not value: return ""
        text = str(value).strip()
        text = re.sub(r'\s*\(.*?\)', '', text)
        if ',' in text: text = text.split(',')[0]
        text = re.sub(r'^\d+\s+', '', text)
        return text.strip().title()

    def _sort_data(self, df):
        if 'Attendee name: Family name' in df.columns:
            df['_s'] = df['Attendee name: Family name'].astype(str).str.lower()
            df = df.sort_values(by='_s').drop(columns=['_s'])
        return df

    def _save_excel(self, df, path):
        df.to_excel(path, index=False)
        try:
            wb = load_workbook(path)
            ws = wb.active
            ws.freeze_panes = 'A2'

            bold_font = Font(bold=True)
            empty_fill = PatternFill(start_color=AppConfig.EMPTY_CELL_COLOR, end_color=AppConfig.EMPTY_CELL_COLOR, fill_type="solid")

            for cell in ws[1]:
                cell.font = bold_font

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    val = cell.value
                    if val is None or str(val).strip() == "":
                        cell.fill = empty_fill

            for col in ws.columns:
                ws.column_dimensions[get_column_letter(col[0].column)].width = 22

            wb.save(path)
        except Exception as e:
            self.log(f"Warnung beim Stylen: {e}")